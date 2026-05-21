from __future__ import annotations

import csv
import json
import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook


def clean_row(values: tuple[object, ...]) -> list[object | None]:
    return [value for value in values]


def image_payloads(output_dir: Path, sheet) -> list[dict[str, object | None]]:
    images_dir = output_dir / "images"
    images_dir.mkdir(exist_ok=True)
    payloads = []

    for index, image in enumerate(getattr(sheet, "_images", []), start=1):
        image_ref = getattr(image, "ref", None)
        image_bytes = None
        if hasattr(image_ref, "read"):
            image_bytes = image_ref.read()
        elif isinstance(image_ref, bytes):
            image_bytes = image_ref

        image_format = getattr(image, "format", None) or "png"
        image_name = f"{sheet.title}_image_{index}.{str(image_format).lower()}"
        image_path = images_dir / image_name

        if image_bytes is not None:
            image_path.write_bytes(image_bytes)
        else:
            image_data = image._data()
            if isinstance(image_data, bytes):
                image_path.write_bytes(image_data)
            else:
                image_path.write_bytes(BytesIO(image_data).getvalue())

        anchor = getattr(image, "anchor", None)
        payloads.append(
            {
                "file": str(image_path.relative_to(output_dir)),
                "width": getattr(image, "width", None),
                "height": getattr(image, "height", None),
                "format": image_format,
                "anchor": getattr(anchor, "_from", None) and {
                    "row": anchor._from.row + 1,
                    "column": anchor._from.col + 1,
                    "cell": f"{chr(anchor._from.col + 65)}{anchor._from.row + 1}",
                },
            }
        )

    return payloads


def sheet_payload(output_dir: Path, sheet) -> dict[str, object]:
    rows = [clean_row(row) for row in sheet.iter_rows(values_only=True)]
    non_empty_rows = [row for row in rows if any(value not in (None, "") for value in row)]
    return {
        "title": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "non_empty_row_count": len(non_empty_rows),
        "image_count": len(getattr(sheet, "_images", [])),
        "images": image_payloads(output_dir, sheet),
        "rows": rows,
        "preview": non_empty_rows[:5],
    }


def write_csv(output_dir: Path, sheet_name: str, rows: list[list[object | None]]) -> None:
    safe_name = "".join(char if char.isalnum() or char in ("-", "_") else "_" for char in sheet_name).strip("_")
    target = output_dir / f"{safe_name or 'sheet'}.csv"
    with target.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)


def extract_workbook(workbook_path: Path) -> Path:
    workbook = load_workbook(workbook_path, data_only=True)
    output_dir = workbook_path.parent / f"{workbook_path.stem}_extract"
    output_dir.mkdir(exist_ok=True)

    payload = {
        "source_file": workbook_path.name,
        "sheet_names": workbook.sheetnames,
        "sheets": [],
    }

    for sheet in workbook.worksheets:
        sheet_data = sheet_payload(output_dir, sheet)
        payload["sheets"].append(sheet_data)
        write_csv(output_dir, sheet.title, sheet_data["rows"])

    json_path = output_dir / "workbook_data.json"
    json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return json_path


def main() -> int:
    if len(sys.argv) > 1:
        workbook_path = Path(sys.argv[1]).expanduser().resolve()
    else:
        workbook_path = Path("2026 Symposium List (version 2).xlsx").resolve()

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    json_path = extract_workbook(workbook_path)
    print(f"Extracted workbook data to: {json_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())